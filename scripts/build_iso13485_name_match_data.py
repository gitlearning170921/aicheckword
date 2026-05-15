"""
从《医疗器械质量管理体系文件清单》xlsx 抽取「中文文件名称 ↔ 英文名称」对，
生成供 aiword 初稿页模板预匹配使用的 JSON（可提交入库；无 xlsx 时跳过）。

默认 xlsx 路径可通过环境变量 ISO13485_FILE_LIST_XLSX 覆盖。
"""
from __future__ import annotations

import json
import os
import re
import sys
from pathlib import Path

try:
    import openpyxl
except ImportError as e:
    print("需要 openpyxl:", e, file=sys.stderr)
    sys.exit(1)

DEFAULT_XLSX = Path(
    r"g:\互联网产品部\质量体系\注册体系文件\【02】管理体系\【01】ISO13485\医疗器械质量管理体系文件清单.xlsx"
)

# aiword 与 aicheckword 常见为同级目录
REPO_ROOT = Path(__file__).resolve().parents[1]
AIWORD_STATIC = REPO_ROOT.parent / "aiword" / "web" / "static" / "data"
OUT_JSON = AIWORD_STATIC / "iso13485_document_name_pairs.json"

# 表头别名（归一化后匹配）
CN_HEADERS = ("文件名称", "文件名", "中文名称", "名称")
EN_HEADERS = ("英文名称", "英文名", "英文")


def _norm_header(cell: object) -> str:
    if cell is None:
        return ""
    s = str(cell).strip().replace("\n", "").replace(" ", "")
    return s


def _find_col(header_row: tuple[object, ...], candidates: tuple[str, ...]) -> int | None:
    for i, cell in enumerate(header_row):
        h = _norm_header(cell)
        for c in candidates:
            if c in h or h in c:
                return i
    return None


def _clean(s: object) -> str:
    if s is None:
        return ""
    t = str(s).strip()
    t = re.sub(r"\s+", " ", t)
    return t


def extract_pairs(path: Path) -> list[dict[str, str]]:
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    pairs: dict[tuple[str, str], None] = {}
    for sn in wb.sheetnames:
        ws = wb[sn]
        it = ws.iter_rows(values_only=True)
        try:
            header = next(it)
        except StopIteration:
            continue
        if not header:
            continue
        icn = _find_col(header, CN_HEADERS)
        ien = _find_col(header, EN_HEADERS)
        if icn is None or ien is None:
            continue
        for row in it:
            if not row or max(icn, ien) >= len(row):
                continue
            cn = _clean(row[icn])
            en = _clean(row[ien])
            if len(cn) < 2 or len(en) < 2:
                continue
            key = (cn, en)
            if key not in pairs:
                pairs[key] = None
    wb.close()
    out = [{"zh": a, "en": b} for (a, b) in pairs.keys()]
    out.sort(key=lambda x: (x["zh"], x["en"]))
    return out


def main() -> int:
    src = Path(os.environ.get("ISO13485_FILE_LIST_XLSX", str(DEFAULT_XLSX)))
    if not src.is_file():
        print("skip: xlsx not found:", src, file=sys.stderr)
        return 0
    data = extract_pairs(src)
    meta = {
        "source": str(src),
        "pair_count": len(data),
        "pairs": data,
    }
    AIWORD_STATIC.mkdir(parents=True, exist_ok=True)
    OUT_JSON.write_text(json.dumps(meta, ensure_ascii=False, indent=2), encoding="utf-8")
    print("wrote", OUT_JSON, "pairs", len(data))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
